from orm import ORM
from sale import Sale
from purchaser import Purchaser
from bid import Bid
from config import DB
from openpyxl import load_workbook
import os


def create_tables():
    if os.path.exists(DB):
        os.remove(DB)
    ORM.create_tables(DB)


def display_sales_extract_dict(sales_dict):
    for sale in sales_dict:
        print(sale)
        for sub in sales_dict[sale]:
            print(f'\t{sub}: {sales_dict[sale][sub]}')
        print()


# Extract from Excel Sheet
def extract():
    excel = 'data/p_tend.xlsx'
    wb = load_workbook(excel, data_only=True)
    ws = wb['Sale Auction Data']

    sales = {}
    purchasers = []
    bids = {}

    for z, i in enumerate(ws.iter_rows()):
        if z != 0:
            sale_name = i[3].value
            if sale_name is None:
                break
            else:
                if sale_name != 'Crow Bait':
                    fy = i[2].value
                    qtr = i[1].value
                    month = i[0].value
                    volume = i[4].value
                    min_bid = i[5].value

                    purch_name = i[11].value
                    if i[12].value.upper() == 'Y':
                        win = 1
                    else:
                        win = 0
                    purch_bid = i[6].value
                    if purch_name is not None:
                        purchasers.append(purch_name)

                        if sale_name not in sales:
                            sales[sale_name] = {
                                'name': sale_name,
                                'fy': fy,
                                'qtr': qtr,
                                'month': month,
                                'volume': volume,
                                'min_bid': min_bid,
                                'min_bid_mbf': min_bid / volume
                            }
                            bids[sale_name] = {'purchasers': {purch_name: {'bid': purch_bid,
                                                                           'win': win}}}
                        else:
                            bids[sale_name]['purchasers'].update({purch_name: {'bid': purch_bid,
                                                                               'win': win}})
    return sales, purchasers, bids


# Transform to database model
def transform(sales, purchasers, bids):
    purchasers = {i for i in purchasers}

    purchasers_ref = {}
    for purchaser_name in purchasers:
        purchaser = Purchaser(name=purchaser_name)
        purchaser.set_db(DB)
        purchaser.insert_self()
        purchasers_ref[purchaser_name] = purchaser.get_last_primary(purchaser.db, Purchaser)

    sales_ref = {}
    for sale_name in sales:
        sale = Sale(**sales[sale_name])
        sale.set_db(DB)
        sale.insert_self()
        sales_ref[sale_name] = sale.get_last_primary(sale.db, Sale)

    for sale_name in bids:
        for purch_name in bids[sale_name]['purchasers']:
            bid = Bid(sale_ref=sales_ref[sale_name],
                      purchaser_ref=purchasers_ref[purch_name],
                      bid=bids[sale_name]['purchasers'][purch_name]['bid'],
                      win=bids[sale_name]['purchasers'][purch_name]['win'],)
            bid.set_db(DB)
            bid.insert_self()


if __name__ == '__main__':
    create_tables()
    transform(*extract())


